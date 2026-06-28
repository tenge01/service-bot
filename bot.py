import logging
import os
import asyncio
import io
from datetime import datetime
from aiogram import Bot, Dispatcher, types, F
from aiogram.filters import CommandStart
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.types import ReplyKeyboardMarkup, KeyboardButton, ReplyKeyboardRemove, InlineKeyboardMarkup, InlineKeyboardButton
import openpyxl
import psycopg2

BOT_TOKEN = os.getenv("BOT_TOKEN", "")
ADMIN_CHAT_ID = int(os.getenv("ADMIN_CHAT_ID", "0"))
DATABASE_URL = os.getenv("DATABASE_URL", "")
MANAGER_PHONE = "+77751340961"
MANAGER_NAME = "Xfinity Solutions"

logging.basicConfig(level=logging.INFO)
log = logging.getLogger(__name__)

bot = Bot(token=BOT_TOKEN)
dp = Dispatcher(storage=MemoryStorage())

TEXTS = {
    "ru": {
        "welcome": "Здравствуйте! Это сервис подачи заявок на ремонт оборудования.\n\nУкажите ваш номер телефона:",
        "ask_city": "Из какого вы города?",
        "ask_school": "Укажите название вашей школы:",
        "ask_serial": "Отправьте серийный номер оборудования.\n\nЕго можно найти на наклейке сзади или снизу устройства.",
        "ask_description": "Опишите неисправность. Что случилось с оборудованием?",
        "check": "Проверьте данные заявки:\n\nТелефон: {phone}\nГород: {city}\nШкола: {school}\nСерийный номер: {serial}\nНеисправность: {description}\n\nВсё верно?",
        "btn_confirm": "Все верно, отправить",
        "btn_restart": "Начать заново",
        "accepted": "Заявка принята!\n\nНомер вашей заявки: {ticket_num}\n\nСервисный центр свяжется с вами в ближайшее время.\n\nЕсли появятся вопросы, обращайтесь к менеджеру {manager}:\n{phone}",
        "restarting": "Хорошо, начнём сначала.\n\nУкажите ваш номер телефона:",
        "invalid_phone": "Пожалуйста, введите корректный номер телефона:",
        "ready": "Ваше оборудование готово к выдаче!\n\nНомер заявки: {ticket_num}\n\nПо вопросам обращайтесь к менеджеру {manager}:\n{phone}",
    },
    "kz": {
        "welcome": "Салеметсіз бе! Бул жабдыкты жондеуге отінім беру кызметі.\n\nТелефон номіріңізді енгізіңіз:",
        "ask_city": "Сіз кай каладансыз?",
        "ask_school": "Мектебіңіздің атауын енгізіңіз:",
        "ask_serial": "Жабдыктың сериялык номірін жіберіңіз.\n\nОны курылгының артky немесе томенгі жагындагы жапсырмадан таба аласыз.",
        "ask_description": "Акауды сипаттаңыз. Жабдыкка не болды?",
        "check": "Отінім маліметтерін тексеріңіз:\n\nТелефон: {phone}\nКала: {city}\nМектеп: {school}\nСериялык номір: {serial}\nАкау: {description}\n\nБарлыгы дурыс па?",
        "btn_confirm": "Барлыгы дурыс, жіберу",
        "btn_restart": "Кайта бастау",
        "accepted": "Отінім кабылданды!\n\nОтінімініздің номірі: {ticket_num}\n\nКызмет корсету орталыгы сізбен жакын арада хабарласады.\n\nСурактарыңыз болса, {manager} менеджеріне хабарласыңыз:\n{phone}",
        "restarting": "Жарайды, кайта бастайык.\n\nТелефон номіріңізді енгізіңіз:",
        "invalid_phone": "Телефон номірін дурыс енгізіңіз:",
        "ready": "Сіздің жабдыгыңыз беруге дайын!\n\nОтінім номірі: {ticket_num}\n\nСурактар бойынша {manager} менеджеріне хабарласыңыз:\n{phone}",
    }
}

def t(lang, key, **kwargs):
    text = TEXTS.get(lang, TEXTS["ru"]).get(key, "")
    return text.format(**kwargs) if kwargs else text

class Form(StatesGroup):
    language    = State()
    phone       = State()
    city        = State()
    school      = State()
    serial      = State()
    description = State()
    confirm     = State()

def get_conn():
    return psycopg2.connect(DATABASE_URL)

def init_db():
    con = get_conn()
    cur = con.cursor()
    cur.execute("""
        CREATE TABLE IF NOT EXISTS tickets (
            id              SERIAL PRIMARY KEY,
            ticket_num      TEXT,
            phone           TEXT,
            city            TEXT,
            school          TEXT,
            serial          TEXT,
            description     TEXT,
            service         TEXT,
            service_contact TEXT,
            status          TEXT DEFAULT 'новая',
            created_at      TEXT,
            tg_user_id      BIGINT
        )
    """)
    con.commit()
    cur.close()
    con.close()

def save_ticket(data: dict):
    con = get_conn()
    cur = con.cursor()
    cur.execute("""
        INSERT INTO tickets
        (ticket_num, phone, city, school, serial, description, service, service_contact, created_at, tg_user_id)
        VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
    """, (
        data["ticket_num"], data["phone"], data["city"], data["school"],
        data["serial"], data["description"], data["service"],
        data["service_contact"], data["created_at"], data["tg_user_id"]
    ))
    con.commit()
    cur.close()
    con.close()

def get_ticket_by_num(ticket_num: str):
    con = get_conn()
    cur = con.cursor()
    cur.execute("SELECT * FROM tickets WHERE ticket_num=%s", (ticket_num,))
    row = cur.fetchone()
    cur.close()
    con.close()
    return row

def get_all_tickets():
    con = get_conn()
    cur = con.cursor()
    cur.execute("SELECT * FROM tickets ORDER BY id DESC")
    rows = cur.fetchall()
    cur.close()
    con.close()
    return rows

def update_status(ticket_num: str, status: str):
    con = get_conn()
    cur = con.cursor()
    cur.execute("UPDATE tickets SET status=%s WHERE ticket_num=%s", (status, ticket_num))
    con.commit()
    cur.close()
    con.close()

def gen_ticket_num() -> str:
    now = datetime.now()
    con = get_conn()
    cur = con.cursor()
    cur.execute("SELECT COUNT(*) FROM tickets")
    count = cur.fetchone()[0] + 1
    cur.close()
    con.close()
    return f"ZV-{now.strftime('%Y%m%d')}-{count:04d}"

def find_service(city: str) -> dict:
    path = "data/services.xlsx"
    if not os.path.exists(path):
        return {"service": "Не определён", "contact": "Не указан"}
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    city_lower = city.strip().lower()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        row_city = str(row[0]).strip().lower()
        if row_city in city_lower or city_lower in row_city:
            return {
                "service": str(row[3]) if row[3] else "Не указан",
                "contact": str(row[4]) if row[4] else "Не указан"
            }
    return {"service": f"Сервис для города {city} не найден", "contact": "Уточните вручную"}

def make_ready_keyboard(ticket_num: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup(inline_keyboard=[[
        InlineKeyboardButton(
            text="Оборудование готово - уведомить школу",
            callback_data=f"ready:{ticket_num}"
        )
    ]])

# ─── Хэндлеры диалога ────────────────────────────────────────

@dp.message(CommandStart())
async def cmd_start(message: types.Message, state: FSMContext):
    await state.clear()
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text="Русский")],
            [KeyboardButton(text="Казахский / Казакша")]
        ],
        resize_keyboard=True
    )
    await message.answer("Выберите язык / Тілді таңдаңыз:", reply_markup=kb)
    await state.set_state(Form.language)

@dp.message(Form.language)
async def got_language(message: types.Message, state: FSMContext):
    if "азак" in message.text.lower() or "азак" in message.text.lower():
        lang = "kz"
    else:
        lang = "ru"
    await state.update_data(lang=lang)
    await message.answer(t(lang, "welcome"), reply_markup=ReplyKeyboardRemove())
    await state.set_state(Form.phone)

@dp.message(Form.phone)
async def got_phone(message: types.Message, state: FSMContext):
    data = await state.get_data()
    lang = data.get("lang", "ru")
    if len(message.text.strip()) < 7:
        await message.answer(t(lang, "invalid_phone"))
        return
    await state.update_data(phone=message.text.strip())
    await message.answer(t(lang, "ask_city"))
    await state.set_state(Form.city)

@dp.message(Form.city)
async def got_city(message: types.Message, state: FSMContext):
    data = await state.get_data()
    await state.update_data(city=message.text.strip())
    await message.answer(t(data.get("lang", "ru"), "ask_school"))
    await state.set_state(Form.school)

@dp.message(Form.school)
async def got_school(message: types.Message, state: FSMContext):
    data = await state.get_data()
    await state.update_data(school=message.text.strip())
    await message.answer(t(data.get("lang", "ru"), "ask_serial"))
    await state.set_state(Form.serial)

@dp.message(Form.serial)
async def got_serial(message: types.Message, state: FSMContext):
    data = await state.get_data()
    await state.update_data(serial=message.text.strip())
    await message.answer(t(data.get("lang", "ru"), "ask_description"))
    await state.set_state(Form.description)

@dp.message(Form.description)
async def got_description(message: types.Message, state: FSMContext):
    await state.update_data(description=message.text.strip())
    data = await state.get_data()
    lang = data.get("lang", "ru")
    kb = ReplyKeyboardMarkup(
        keyboard=[
            [KeyboardButton(text=t(lang, "btn_confirm"))],
            [KeyboardButton(text=t(lang, "btn_restart"))]
        ],
        resize_keyboard=True
    )
    await message.answer(
        t(lang, "check",
          phone=data["phone"], city=data["city"],
          school=data["school"], serial=data["serial"],
          description=data["description"]),
        reply_markup=kb
    )
    await state.set_state(Form.confirm)

@dp.message(Form.confirm)
async def got_confirm(message: types.Message, state: FSMContext):
    data = await state.get_data()
    lang = data.get("lang", "ru")

    if message.text == t(lang, "btn_restart"):
        await state.clear()
        await message.answer(t(lang, "restarting"), reply_markup=ReplyKeyboardRemove())
        await state.set_state(Form.phone)
        return

    if message.text != t(lang, "btn_confirm"):
        return

    ticket_num = gen_ticket_num()
    created_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    service_info = find_service(data["city"])

    full_data = {
        **data,
        "ticket_num": ticket_num,
        "created_at": created_at,
        "service": service_info["service"],
        "service_contact": service_info["contact"],
        "tg_user_id": message.from_user.id
    }
    save_ticket(full_data)

    # Ответ школе
    await message.answer(
        t(lang, "accepted",
          ticket_num=ticket_num,
          manager=MANAGER_NAME,
          phone=MANAGER_PHONE),
        reply_markup=ReplyKeyboardRemove()
    )

    # Уведомление администратору с кнопкой
    admin_text = (
        f"НОВАЯ ЗАЯВКА: {ticket_num}\n\n"
        f"Телефон: {full_data['phone']}\n"
        f"Город: {full_data['city']}\n"
        f"Школа: {full_data['school']}\n"
        f"Серийный номер: {full_data['serial']}\n"
        f"Неисправность: {full_data['description']}\n\n"
        f"--- КУДА ОБРАЩАТЬСЯ ---\n"
        f"Сервисный центр: {service_info['service']}\n"
        f"Контакт сервиса: {service_info['contact']}"
    )
    try:
        await bot.send_message(
            ADMIN_CHAT_ID,
            admin_text,
            reply_markup=make_ready_keyboard(ticket_num)
        )
    except Exception as e:
        log.error(f"Ошибка отправки админу: {e}")

    await state.clear()

# ─── Кнопка "Оборудование готово" ────────────────────────────

@dp.callback_query(F.data.startswith("ready:"))
async def equipment_ready(callback: types.CallbackQuery):
    if callback.from_user.id != ADMIN_CHAT_ID:
        return

    ticket_num = callback.data.split(":")[1]
    row = get_ticket_by_num(ticket_num)

    if not row:
        await callback.answer("Заявка не найдена!")
        return

    tg_user_id = row[11]
    update_status(ticket_num, "готово")

    # Сообщение школе на двух языках
    msg_ru = t("ru", "ready", ticket_num=ticket_num, manager=MANAGER_NAME, phone=MANAGER_PHONE)
    msg_kz = t("kz", "ready", ticket_num=ticket_num, manager=MANAGER_NAME, phone=MANAGER_PHONE)
    full_msg = f"{msg_ru}\n\n---\n\n{msg_kz}"

    try:
        await bot.send_message(tg_user_id, full_msg)
        await callback.message.edit_text(
            callback.message.text + "\n\nШкола уведомлена об готовности оборудования!"
        )
        await callback.answer("Школа уведомлена!")
    except Exception as e:
        log.error(f"Ошибка уведомления школы: {e}")
        await callback.answer("Ошибка при отправке!")

# ─── Экспорт в Excel ─────────────────────────────────────────

@dp.message(F.text == "/export")
async def export_excel(message: types.Message):
    if message.from_user.id != ADMIN_CHAT_ID:
        return

    rows = get_all_tickets()
    if not rows:
        await message.answer("Заявок пока нет.")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Заявки"

    headers = ["ID", "Номер заявки", "Телефон", "Город", "Школа",
               "Серийный номер", "Неисправность", "Сервис", "Контакт сервиса",
               "Статус", "Дата создания", "Telegram ID"]
    ws.append(headers)

    for row in rows:
        ws.append(list(row))

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 20

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"tickets_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    await message.answer_document(
        types.BufferedInputFile(buffer.read(), filename=filename),
        caption=f"Все заявки — {len(rows)} шт."
    )

# ─── Команды администратора ───────────────────────────────────

@dp.message(F.text == "/tickets")
async def admin_tickets(message: types.Message):
    if message.from_user.id != ADMIN_CHAT_ID:
        return
    rows = get_all_tickets()
    if not rows:
        await message.answer("Заявок пока нет.")
        return
    text = "Последние заявки:\n\n"
    for r in rows[:10]:
        text += f"{r[1]} | {r[4]} | {r[5]} | {r[9]}\n"
    await message.answer(text)

@dp.message(F.text.startswith("/status "))
async def admin_set_status(message: types.Message):
    if message.from_user.id != ADMIN_CHAT_ID:
        return
    parts = message.text.split(" ", 2)
    if len(parts) < 3:
        await message.answer("Формат: /status ZV-20260629-0001 готово")
        return
    update_status(parts[1], parts[2])
    await message.answer(f"Статус заявки {parts[1]} обновлен: {parts[2]}")

# ─── Запуск ───────────────────────────────────────────────────

async def main():
    init_db()
    log.info("Бот запущен")
    await dp.start_polling(bot)

if __name__ == "__main__":
    asyncio.run(main())
